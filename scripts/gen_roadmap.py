import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

def border(style="thin"):
    s = Side(border_style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def fnt(bold=False, color="000000", size=11, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin   = border("thin")
thick  = border("medium")

NAVY="1B3A5C"; BLUE="2E75B6"; GREEN="375623"; ORANGE="C55A11"
PURPLE="7030A0"; GREY="595959"; LGREY="F2F2F2"; WHITE="FFFFFF"; YELLOW="FFF2CC"

# ── SHEET 1: Gantt Timeline ──────────────────────────────────────────────────
ws = wb.active
ws.title = "Roadmap Timeline"

phases = [
    ("Phase 1\nFoundation",  "Leave Management (apply / approve / balance)",        "Critical", 3, "All industries",                       1,  3),
    ("Phase 1\nFoundation",  "ESS Portal — leave + payslip self-service",           "Critical", 2, "All industries",                       3,  2),
    ("Phase 1\nFoundation",  "Configurable module labels / terminology",             "High",     3, "Non-manufacturing tenants",            4,  3),
    ("Phase 1\nFoundation",  "Custom fields on invoices & items",                   "High",     2, "All industries",                       6,  2),
    ("Phase 2\nService Biz", "Service billing (time+material / milestone)",         "Critical", 3, "IT, consulting, agencies",             8,  3),
    ("Phase 2\nService Biz", "Recurring invoices (AMC / SaaS / rent)",              "High",     2, "SaaS, AMC, rentals",                  10,  2),
    ("Phase 2\nService Biz", "Proforma to Invoice workflow",                        "Medium",   1, "Exporters, traders",                  11,  1),
    ("Phase 2\nService Biz", "Service invoice with TDS deduction",                  "High",     2, "Consulting, agencies",                11,  2),
    ("Phase 2\nService Biz", "Expense reimbursement claims",                        "Medium",   2, "Any company with field staff",        13,  2),
    ("Phase 3\nInventory+",  "Multi-location / multi-warehouse inventory",          "High",     4, "Retail, distribution, construction",  15,  4),
    ("Phase 3\nInventory+",  "Serial number & lot tracking",                        "Medium",   3, "Electronics, pharma, auto parts",     18,  3),
    ("Phase 3\nInventory+",  "Unit of measure conversion",                          "Medium",   2, "Pharma, textiles, chemicals",         20,  2),
    ("Phase 4\nProject+",    "Project cost centre (P&L by project)",                "High",     5, "Construction, consulting, events",    22,  5),
    ("Phase 4\nProject+",    "Progress billing / milestone billing",                "High",     3, "Construction, IT projects",           26,  3),
    ("Phase 4\nProject+",    "Timesheet & billable hours",                          "Medium",   3, "IT, consulting, agencies",            28,  3),
    ("Phase 4\nProject+",    "BOQ (Bill of Quantities)",                            "Medium",   2, "Construction",                        30,  2),
    ("Phase 5\nAdvanced",    "Multi-currency (buy/sell in foreign currency)",        "Medium",   4, "Exporters, importers",                32,  4),
    ("Phase 5\nAdvanced",    "Fixed asset management + depreciation",               "Medium",   3, "Construction, hospitality, mfg",      35,  3),
    ("Phase 5\nAdvanced",    "Performance appraisal module",                        "Low",      3, "All industries",                      37,  3),
    ("Phase 5\nAdvanced",    "Retention management (subcontractor payments)",       "Low",      2, "Construction",                        39,  2),
]

PRIO_FILL = {"Critical": fill("C00000"), "High": fill("C55A11"), "Medium": fill("2E75B6"), "Low": fill(GREY)}
PRIO_FONT = {k: fnt(bold=True, color=WHITE) for k in PRIO_FILL}
PHASE_FILL = {
    "Phase 1\nFoundation":  fill(NAVY),
    "Phase 2\nService Biz": fill(GREEN),
    "Phase 3\nInventory+":  fill(PURPLE),
    "Phase 4\nProject+":    fill(ORANGE),
    "Phase 5\nAdvanced":    fill(GREY),
}
TOTAL_WEEKS = 40

ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 36
ws.row_dimensions[3].height = 16

# Title
ws.merge_cells("A1:F1")
c = ws["A1"]
c.value = "SwachERP — Generic ERP Roadmap Timeline"
c.font = fnt(bold=True, size=14, color=WHITE)
c.fill = fill(NAVY); c.alignment = center

ws.merge_cells(f"G1:{get_column_letter(6+TOTAL_WEEKS)}1")
c = ws["G1"]
c.value = "WEEK-BY-WEEK GANTT  (each shaded cell = 1 week of work)"
c.font = fnt(bold=True, size=11, color=WHITE)
c.fill = fill(NAVY); c.alignment = center

# Column headers
col_headers = ["Phase", "Feature / Deliverable", "Priority", "Effort\n(wks)", "Target Industries", "Status"]
col_widths   = [18,      45,                       10,          8,              34,                   11]
for i, (h, w) in enumerate(zip(col_headers, col_widths), 1):
    c = ws.cell(row=2, column=i, value=h)
    c.font = fnt(bold=True, color=WHITE, size=10)
    c.fill = fill(BLUE); c.alignment = center; c.border = thin
    ws.column_dimensions[get_column_letter(i)].width = w

for w in range(1, TOTAL_WEEKS+1):
    col = 6 + w
    c = ws.cell(row=2, column=col, value=w)
    c.font = fnt(bold=True, color=WHITE, size=8)
    c.fill = fill(BLUE); c.alignment = center; c.border = thin
    ws.column_dimensions[get_column_letter(col)].width = 2.8

# Month labels row 3
month_map = {1:"M1",5:"M2",9:"M3",13:"M4",17:"M5",21:"M6",25:"M7",29:"M8",33:"M9",37:"M10"}
for sw, label in month_map.items():
    ew = min(sw+3, TOTAL_WEEKS)
    ws.merge_cells(start_row=3, start_column=6+sw, end_row=3, end_column=6+ew)
    c = ws.cell(row=3, column=6+sw, value=label)
    c.font = fnt(bold=True, color=WHITE, size=8)
    c.fill = fill(GREY); c.alignment = center; c.border = thin
for ci in range(1, 7):
    ws.cell(row=3, column=ci).fill = fill(LGREY)
    ws.cell(row=3, column=ci).border = thin

# Data rows
prev_phase = None
for ro, (phase, feat, prio, effort, ind, sw, dur) in enumerate(phases):
    r = 4 + ro
    ws.row_dimensions[r].height = 22
    pf = PHASE_FILL[phase]

    c = ws.cell(row=r, column=1, value=phase if phase != prev_phase else "")
    c.font = fnt(bold=(phase != prev_phase), color=WHITE, size=9)
    c.fill = pf; c.alignment = center; c.border = thin
    prev_phase = phase

    c = ws.cell(row=r, column=2, value=feat)
    c.font = fnt(size=9); c.alignment = left; c.border = thin

    c = ws.cell(row=r, column=3, value=prio)
    c.font = PRIO_FONT[prio]; c.fill = PRIO_FILL[prio]; c.alignment = center; c.border = thin

    c = ws.cell(row=r, column=4, value=effort)
    c.font = fnt(bold=True, size=9); c.fill = fill(YELLOW); c.alignment = center; c.border = thin

    c = ws.cell(row=r, column=5, value=ind)
    c.font = fnt(size=9, italic=True); c.alignment = left; c.border = thin

    c = ws.cell(row=r, column=6, value="Planned")
    c.font = fnt(size=9, color=GREY); c.alignment = center; c.border = thin

    for w in range(1, TOTAL_WEEKS+1):
        c = ws.cell(row=r, column=6+w)
        c.border = thin
        c.fill = pf if sw <= w < sw+dur else fill(WHITE)

ws.freeze_panes = "G4"

# ── SHEET 2: Phase Summary ───────────────────────────────────────────────────
ws2 = wb.create_sheet("Phase Summary")

ws2.merge_cells("A1:G1")
c = ws2["A1"]
c.value = "SwachERP — Phase Summary & Business Case"
c.font = fnt(bold=True, size=14, color=WHITE)
c.fill = fill(NAVY); c.alignment = center
ws2.row_dimensions[1].height = 22

hdrs2 = ["Phase", "Timeline", "Features", "Total Effort", "Key Industries Unlocked", "Business Impact", "Target"]
wids2 = [24, 14, 10, 13, 42, 38, 12]
for i, (h, w) in enumerate(zip(hdrs2, wids2), 1):
    c = ws2.cell(row=2, column=i, value=h)
    c.font = fnt(bold=True, color=WHITE, size=10)
    c.fill = fill(BLUE); c.alignment = center; c.border = thin
    ws2.column_dimensions[get_column_letter(i)].width = w

summary = [
    ("Phase 1 — Foundation",  "Month 1-4",  4, "10 weeks", "All industries",                                "Removes manufacturing-only lock-in. Every new vertical becomes sellable.", "Q3 2026"),
    ("Phase 2 — Service Biz", "Month 3-6",  5, "10 weeks", "IT, consulting, AMC, agencies",                 "Unlocks ~40% of Indian GST filers who are pure service companies.",         "Q3 2026"),
    ("Phase 3 — Inventory+",  "Month 4-6",  3, "9 weeks",  "Retail, distribution, pharma, auto parts",      "Multi-branch clients: potential 2x revenue per account.",                   "Q4 2026"),
    ("Phase 4 — Projects",    "Month 6-9",  4, "13 weeks", "Construction, IT projects, events, consulting",  "Construction alone: Rs 12 lakh crore sector. Opens premium project tier.",  "Q1 2027"),
    ("Phase 5 — Advanced",    "Month 8-10", 4, "12 weeks", "Exporters, large enterprises, all industries",   "Enterprise ACV uplift. Multi-currency + assets = large company readiness.", "Q1 2027"),
]
pf_list = [fill(NAVY), fill(GREEN), fill(PURPLE), fill(ORANGE), fill(GREY)]

for ri, (row, pf) in enumerate(zip(summary, pf_list), 3):
    ws2.row_dimensions[ri].height = 36
    for ci, val in enumerate(row, 1):
        c = ws2.cell(row=ri, column=ci, value=val)
        c.border = thin
        c.alignment = Alignment(horizontal="center" if ci in (2,3,4,7) else "left", vertical="center", wrap_text=True)
        if ci == 1:
            c.font = fnt(bold=True, color=WHITE, size=10); c.fill = pf
        elif ci == 4:
            c.font = fnt(bold=True, size=10); c.fill = fill(YELLOW)
        elif ci == 7:
            c.font = fnt(italic=True, color=GREY, size=9)
        else:
            c.font = fnt(size=9)

rt = len(summary) + 3
ws2.row_dimensions[rt].height = 22
ws2.merge_cells(f"A{rt}:B{rt}")
for ci in range(1, 8):
    c = ws2.cell(row=rt, column=ci)
    c.fill = fill(NAVY); c.border = thick
    c.font = fnt(bold=True, color=WHITE, size=11)
    c.alignment = center
ws2.cell(row=rt, column=1).value = "TOTAL ROADMAP"
ws2.cell(row=rt, column=3).value = sum(s[2] for s in summary)
ws2.cell(row=rt, column=4).value = "~54 weeks"

# ── SHEET 3: Gap Analysis ────────────────────────────────────────────────────
ws3 = wb.create_sheet("Gap Analysis")

ws3.merge_cells("A1:G1")
c = ws3["A1"]
c.value = "SwachERP — Detailed Gap Analysis: Manufacturing vs Generic"
c.font = fnt(bold=True, size=14, color=WHITE)
c.fill = fill(NAVY); c.alignment = center
ws3.row_dimensions[1].height = 22

hdrs3 = ["Category", "Gap", "Current State", "What Is Needed", "Priority", "Effort", "Phase"]
wids3 = [18, 36, 30, 32, 10, 9, 18]
for i, (h, w) in enumerate(zip(hdrs3, wids3), 1):
    c = ws3.cell(row=2, column=i, value=h)
    c.font = fnt(bold=True, color=WHITE, size=10)
    c.fill = fill(BLUE); c.alignment = center; c.border = thin
    ws3.column_dimensions[get_column_letter(i)].width = w

CAT_FILL = {
    "Terminology":   fill("BDD7EE"),
    "HR":            fill("E2EFDA"),
    "Billing":       fill("FFF2CC"),
    "Inventory":     fill("EAD1DC"),
    "Projects":      fill("FCE4D6"),
    "Finance":       fill("DAEEF3"),
    "Configuration": fill(LGREY),
}

gaps = [
    ("Terminology",   "Hardcoded Production Orders",         "Fixed label shown to all tenants",          "Configurable per tenant: Work Order, Job, Project",   "Critical","2 wks","Phase 1"),
    ("Terminology",   "Hardcoded BOM",                       "Bill of Materials only",                    "Recipe / Service Package / Scope of Work",            "High",    "1 wk", "Phase 1"),
    ("Terminology",   "Gatepass shown to everyone",          "Delivery challan term baked into navigation","Hide or relabel for non-manufacturing tenants",       "High",    "1 wk", "Phase 1"),
    ("Terminology",   "Machine Startup Checklist visible",   "WhatsApp machine checklist in nav always",  "Show only if tenant has machines / production lines",  "Medium",  "1 wk", "Phase 1"),
    ("HR",            "No leave management",                 "Attendance only. No apply/approve/balance", "Leave types, application, approval, running balance",  "Critical","3 wks","Phase 1"),
    ("HR",            "No expense claims",                   "Not present",                               "Employee submits claim, manager approves, finance pays","Medium", "2 wks","Phase 2"),
    ("HR",            "No timesheet / billable hours",       "Shift attendance only",                     "Daily timesheet linked to project or client",          "Medium",  "3 wks","Phase 4"),
    ("HR",            "No performance appraisal",            "Not present",                               "Appraisal cycles, KRA, ratings, increment link",       "Low",     "3 wks","Phase 5"),
    ("Billing",       "No service billing",                  "Product invoices (HSN) only",               "Service lines: time+material, milestone, retainer",    "Critical","3 wks","Phase 2"),
    ("Billing",       "No recurring invoices",               "Manual invoice creation only",              "Auto-generate monthly/quarterly invoices on schedule", "High",    "2 wks","Phase 2"),
    ("Billing",       "No proforma invoice",                 "Tax Invoice only — no pre-invoice stage",   "Proforma -> convert to Tax Invoice flow",              "Medium",  "1 wk", "Phase 2"),
    ("Billing",       "TDS not on service purchase bills",   "TDS tracking exists, not on bill creation", "Auto-calculate TDS on applicable service vendor bills","High",    "2 wks","Phase 2"),
    ("Billing",       "No progress/milestone billing",       "Full invoice only",                         "Running Account bill, partial billing on milestones",  "High",    "3 wks","Phase 4"),
    ("Inventory",     "Single warehouse only",               "Company-level stock, no site/branch",       "Site/branch/godown-wise stock with stock transfers",   "High",    "4 wks","Phase 3"),
    ("Inventory",     "No serial / lot tracking",            "Batch FIFO only",                           "Per-unit serial tracking and lot traceability",        "Medium",  "3 wks","Phase 3"),
    ("Inventory",     "No UOM conversion",                   "Single unit per item",                      "Buy in kg, sell in grams — convert at transaction",   "Medium",  "2 wks","Phase 3"),
    ("Projects",      "No project cost centre",              "No project entity in system",               "Project master, link POs, invoices, costs to project", "High",    "5 wks","Phase 4"),
    ("Projects",      "No BOQ tracking",                     "Not present",                               "BOQ vs actual material and labour by project",         "Medium",  "2 wks","Phase 4"),
    ("Finance",       "No multi-currency",                   "INR only throughout",                       "Forex invoices, bank accounts, gain/loss accounting",  "Medium",  "4 wks","Phase 5"),
    ("Finance",       "No fixed asset register",             "Not present",                               "Asset master, depreciation schedule, disposal",        "Medium",  "3 wks","Phase 5"),
    ("Finance",       "No retention management",             "Not present",                               "Hold % of subcontractor payment until project done",   "Low",     "2 wks","Phase 5"),
    ("Configuration", "No custom fields",                    "Fixed schema on all forms",                 "User-defined fields on invoices, items, employees",    "High",    "2 wks","Phase 1"),
    ("Configuration", "No doc numbering config",             "System-generated fixed format",             "Tenant configures prefix, year, series (PROJ-24-001)","Medium",  "1 wk", "Phase 1"),
]

prev_cat = None
for ri, row in enumerate(gaps, 3):
    r = ri + 2
    ws3.row_dimensions[r].height = 24
    for ci, val in enumerate(row, 1):
        c = ws3.cell(row=r, column=ci, value=val)
        c.border = thin
        c.alignment = Alignment(horizontal="center" if ci in (1,5,6,7) else "left", vertical="center", wrap_text=True)
        cat = row[0]
        if ci == 1:
            c.fill = CAT_FILL.get(cat, fill(LGREY))
            c.font = fnt(bold=(cat != prev_cat), size=9)
        elif ci == 5:
            c.fill = PRIO_FILL.get(val, fill(LGREY)); c.font = PRIO_FONT.get(val, fnt(size=9))
        else:
            c.font = fnt(size=9)
    prev_cat = row[0]

ws3.freeze_panes = "A3"

wb.save("SwachERP_Generic_Roadmap.xlsx")
print("Saved SwachERP_Generic_Roadmap.xlsx")
